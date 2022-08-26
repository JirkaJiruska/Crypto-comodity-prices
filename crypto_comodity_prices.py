import coinmarketcapapi, sys, os, urllib.request, re
from openpyxl.styles import Font
import openpyxl
import xlsxwriter

#############################################
# CZK/USD EXCHANGE RATE
#############################################
# load CZK/USD exchange rate from:
# open html source code by pressing Ctrl+U at following link: https://www.penize.cz/kurzy-men/6591-americky-dolar
# and find the "Kč/USD" string, which is written in html as "K\\xc4\\x8d/USD"
# previous characters are representing the exchange rate number

link_exchange_rate = "https://www.penize.cz/kurzy-men/6591-americky-dolar"
html = urllib.request.urlopen(link_exchange_rate)
sub_html = str(html.read())
index = sub_html.find(" K\\xc4\\x8d/USD")
CZK_USD_exchange_rate = float(sub_html[index-6: index-4] + "." + sub_html[index-3: index])
print(CZK_USD_exchange_rate, "CZK/USD")

#############################################
# GOLD and SILVER PRICE from AU portal
# GOLD
#############################################
link_au_portal = "https://www.auportal.cz/"
html = urllib.request.urlopen(link_au_portal)
html_lines = html.readlines()
# characteristic set of characters one line above the gold price
line_pattern = '<span>Zlato</span>'
line_index = 1
au_line = ""
au_price_str = ""
au_price = 0
# find the line where the pattern is and one line below there is a gold price
for line in html_lines:
    if line_pattern in str(line):
        line_index = html_lines.index(line) + 1
        au_line = str(html_lines[line_index])
        print("html line for gold: ", au_line)
        # first digit is a frst digit of the price - we are adding digit until we find the comma
        # the line with price looks like this: 
        # b'   <span class="gray">39 417,32 K\xc4\x8d/oz</span>\n'
        for char in au_line:
            if char == ",":
                break
            if char.isdigit():
                au_price_str = au_price_str + char

        au_price_str = str(au_price_str)
        au_price = float(au_price_str)

#############################################
# SILVER
#############################################

link_au_portal = "https://www.auportal.cz/"
html = urllib.request.urlopen(link_au_portal)
html_lines = html.readlines()

# characteristic set of characters one line above the silver price
line_pattern = '<span>St\\xc5\\x99\\xc3\\xadbro</span>'
line_index = 1
ag_line = ""
ag_price_str = ""
ag_price = 0
# find the line where the pattern is and one line below there is a silver price
for line in html_lines:
    if line_pattern in str(line):
        line_index = html_lines.index(line) + 1
        ag_line = str(html_lines[line_index])
        print("html line for silver: ", ag_line)
        # first digit is a frst digit of the price - we are adding digit until we find the comma
        # the line with price looks like this: 
        # b'   <span class="gray">39 417,32 K\xc4\x8d/oz</span>\n'
        for char in ag_line:
            if char == ",":
                break
            if char.isdigit():
                ag_price_str = ag_price_str + char

        ag_price_str = str(ag_price_str)
        ag_price = float(ag_price_str)


#############################################
# CRYPTO
#############################################
# list of interrested crypto
currency_list = ["Bitcoin", "Ethereum", "Litecoin", "Polkadot", "Cardano", "Uniswap", "Chainlink", "Eos", "Dash"]
currency_tickers = {"Bitcoin":"BTC", "Ethereum":"ETH", "Litecoin":"LTC", "Polkadot":"DOT", "Cardano":"ADA", "Uniswap":"UNI", "Chainlink":"LINK", "Eos":"EOS", "Dash":"DASH"}
# get access to coinmarketcap.com
cmc = coinmarketcapapi.CoinMarketCapAPI("01bfa8c2-e86d-45c4-8f11-91ee9a49b349")


home_path = os.path.dirname(sys.argv[0])
excel_path = home_path + "/data_store.xlsx"
if not os.path.exists(excel_path):
    xls_file = xlsxwriter.Workbook(excel_path)
    xls_file.close()
    print("File has been created")


xlx_workbook = openpyxl.load_workbook(excel_path)

head_font = Font(bold=True)

sheet = xlx_workbook.active
headline = sheet['A1']
headline.font = head_font

xlx_workbook.worksheets[0].cell(1, 1, "Asset")
xlx_workbook.worksheets[0].cell(1, 2, "Ticker")
xlx_workbook.worksheets[0].cell(1, 3, "USD")
xlx_workbook.worksheets[0].cell(1, 4, "CZK")


row = 2
for i in currency_list:
    # get data from coinmarketcap.com
    data = cmc.cryptocurrency_quotes_latest(symbol=currency_tickers[i], convert='USD')
    print("{s:10}\t{t}\t{a:5.2f} USD".format(s=i, t=currency_tickers[i], a=data.data[currency_tickers[i]]["quote"]["USD"]["price"]))
    # write crypto prices into excel table
    if len(xlx_workbook.sheetnames) > 0:
        xlx_workbook.worksheets[0].cell(row, 1, i)
        xlx_workbook.worksheets[0].cell(row, 2, currency_tickers[i])
        xlx_workbook.worksheets[0].cell(row, 3, data.data[currency_tickers[i]]["quote"]["USD"]["price"])
        xlx_workbook.worksheets[0].cell(row, 4, data.data[currency_tickers[i]]["quote"]["USD"]["price"] * CZK_USD_exchange_rate)
        row = row + 1 

# write gold price into excel table
xlx_workbook.worksheets[0].cell(row, 1, "Zlato 1oz")
xlx_workbook.worksheets[0].cell(row, 2, "GOLD")
xlx_workbook.worksheets[0].cell(row, 3, au_price / CZK_USD_exchange_rate)
xlx_workbook.worksheets[0].cell(row, 4, au_price)
print("{s:10}\t{t}\t{a:5.2f} USD".format(s="Zlato 1oz", t="GOLD", a=au_price / CZK_USD_exchange_rate))
row = row + 1

# write silver price into excel table
xlx_workbook.worksheets[0].cell(row, 1, "Stříbro 1oz")
xlx_workbook.worksheets[0].cell(row, 2, "SILVER")
xlx_workbook.worksheets[0].cell(row, 3, ag_price / CZK_USD_exchange_rate)
xlx_workbook.worksheets[0].cell(row, 4, ag_price)
print("{s:10}\t{t}\t{a:5.2f} USD".format(s="Stříbro 1oz", t="SILVER", a=ag_price / CZK_USD_exchange_rate))



xlx_workbook.save(excel_path)
xlx_workbook.close()


